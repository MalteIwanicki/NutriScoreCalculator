def _extract_formulars():
    import openpyxl

    workbook = openpyxl.load_workbook(
        "./nutri-score-dt-excel-berechnungstabelle.xlsx",
        read_only=True,
    )
    sheetnames = ["allgemeiner Fall", "Käse", "zugesetzte Fette", "Getränke"]
    sheet_formulars = {}
    for sheetname in sheetnames:
        worksheet = workbook[sheetname]
        header = worksheet[1]
        columns = {col.column_letter: col.value for col in header}
        formulars = []
        for col, name in columns.items():
            cell = col + "2"
            formulars.append(f"{cell}, {name}: {worksheet[cell].value}")
        sheet_formulars[sheetname] = formulars
    workbook.close()
    import json

    with open("formulars.json", "w", encoding="utf-8") as f:
        json.dump(sheet_formulars, f, indent=6, ensure_ascii=False)


if __name__ == "__main__":
    _extract_formulars()
