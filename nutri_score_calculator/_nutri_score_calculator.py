from enum import Enum


def _all_parameter_set(func):
    def wrapper(*args, **kwargs):
        if not all(arg is not None for arg in args) or not all(
            value is not None for value in kwargs.values()
        ):
            return None
        else:
            return func(*args, **kwargs)

    return wrapper


class NurtiScoreGeneral:
    @staticmethod
    def _calculate_brennwert(kilokalorien):
        return round(kilokalorien * 4.184, 0)

    @staticmethod
    def _calculate_natrium(salz):
        return round(salz / 2.5 * 1000, 0)

    @staticmethod
    def _calculate_rating(ist_wasser, wert):
        if wert is None:
            return "fehlende_Daten"
        if wert < 0:
            return "Nutriscore_A"
        elif wert < 3:
            return "Nutriscore_B"
        elif wert < 11:
            return "Nutriscore_C"
        elif wert < 19:
            return "Nutriscore_D"
        else:
            return "Nutriscore_E"

    @classmethod
    def calculate_nutri_score(
        cls,
        kilokalorien,
        gesaettigte_fettsaeuren,
        zucker,
        proteine,
        ballaststoffe,
        anteil_obst_gemuese_huelsen_schalen_raps_walnuss_und_olivenoele,
        salz=None,
        natrium=None,
        gesamtfett=None,
        ist_wasser=False,
    ):
        if salz == None and natrium == None:
            raise Exception("Salz oder Natrium muss angegeben werden")
        elif not salz == None:
            natrium_calculated = cls._calculate_natrium(salz)
            if not natrium == None:
                if natrium_calculated != natrium:
                    raise Exception(
                        "Entweder Salz und Natrium weichen stimmen nicht überein"
                    )
            natrium = natrium_calculated

        brennwert = cls._calculate_brennwert(kilokalorien)
        pts_kj = cls._calculate_pts_kj(brennwert)
        pts_glus = cls._calculate_pts_glus(zucker)
        pts_ags = cls._calculate_pts_ags(gesaettigte_fettsaeuren)
        pts_na = cls._calculate_pts_na(natrium)
        pts_prot = cls._calculate_pts_prot(proteine)
        pts_fib = cls._calculate_pts_fib(ballaststoffe)
        pts_FLN = cls._calculate_pts_FLN(
            anteil_obst_gemuese_huelsen_schalen_raps_walnuss_und_olivenoele
        )
        pts_A = cls._calculate_pts_A(
            pts_kj,
            pts_glus,
            pts_ags,
            pts_na,
        )

        wert = cls._calculate_wert(pts_A, pts_prot, pts_fib, pts_FLN)
        rating = cls._calculate_rating(ist_wasser, wert)
        return rating, wert

    @staticmethod
    @_all_parameter_set
    def _calculate_pts_kj(brennwert):
        wert = round(brennwert, 1)
        thresholds = {
            335: 0,
            670: 1,
            1005: 2,
            1340: 3,
            1675: 4,
            2010: 5,
            2345: 6,
            2680: 7,
            3015: 8,
            3350: 9,
        }
        return next((val for key, val in thresholds.items() if wert <= key), 10)

    @staticmethod
    @_all_parameter_set
    def _calculate_pts_glus(zucker):
        wert = round(zucker, 2)
        thresholds = {
            4.5: 0,
            9: 1,
            13.5: 2,
            18: 3,
            22.5: 4,
            27: 5,
            31: 6,
            36: 7,
            40: 8,
            45: 9,
        }
        return next((val for key, val in thresholds.items() if wert <= key), 10)

    @staticmethod
    @_all_parameter_set
    def _calculate_pts_ags(gesaettigte_fettsaeuren):
        wert = round(gesaettigte_fettsaeuren, 1)
        thresholds = {
            1: 0,
            2: 1,
            3: 2,
            4: 3,
            5: 4,
            6: 5,
            7: 6,
            8: 7,
            9: 8,
            10: 9,
        }
        return next((val for key, val in thresholds.items() if wert <= key), 10)

    @staticmethod
    @_all_parameter_set
    def _calculate_pts_na(natrium):
        wert = round(natrium, 1)
        thresholds = {
            90: 0,
            180: 1,
            270: 2,
            360: 3,
            450: 4,
            540: 5,
            630: 6,
            720: 7,
            810: 8,
            900: 9,
        }
        return next((val for key, val in thresholds.items() if wert <= key), 10)

    @staticmethod
    @_all_parameter_set
    def _calculate_pts_prot(protein):
        wert = round(protein, 2)
        thresholds = {
            1.6: 0,
            3.2: 1,
            4.8: 2,
            6.4: 3,
            8: 4,
        }
        return next((val for key, val in thresholds.items() if wert <= key), 5)

    @staticmethod
    @_all_parameter_set
    def _calculate_pts_fib(ballaststoffe):
        wert = round(ballaststoffe, 2)
        thresholds = {
            0.9: 0,
            1.9: 1,
            2.8: 2,
            3.7: 3,
            4.7: 4,
        }
        for key, val in thresholds.items():
            if wert > key:
                continue
            return val
        return 5

    @staticmethod
    @_all_parameter_set
    def _calculate_pts_FLN(
        anteil_obst_gemuese_huelsen_schalen_raps_walnuss_und_olivenoele,
    ):
        anteile = anteil_obst_gemuese_huelsen_schalen_raps_walnuss_und_olivenoele
        if anteile is None:
            return None
        wert = round(anteil_obst_gemuese_huelsen_schalen_raps_walnuss_und_olivenoele, 1)
        thresholds = {
            40: 0,
            60: 1,
            80: 2,
        }
        return next((val for key, val in thresholds.items() if wert <= key), 5)

    @staticmethod
    @_all_parameter_set
    def _calculate_pts_A(
        pts_kj,
        pts_glus,
        pts_ags,
        pts_na,
    ):
        if any(
            v is None
            for v in [
                pts_kj,
                pts_glus,
                pts_ags,
                pts_na,
            ]
        ):
            return None
        return pts_kj + pts_glus + pts_ags + pts_na

    @staticmethod
    @_all_parameter_set
    def _calculate_wert(pts_A, pts_prot, pts_fib, pts_FLN):
        if 0 <= pts_A < 11:
            return pts_A - sum([pts_prot, pts_fib, pts_FLN])
        elif pts_A >= 11 and pts_FLN == 5:
            return pts_A - sum([pts_prot, pts_fib, pts_FLN])
        else:
            return pts_A - pts_FLN - pts_fib


class Kaese(NurtiScoreGeneral):
    def _calculate_wert(pts_A, pts_prot, pts_fib, pts_FLN):
        if any(v is None for v in [pts_A, pts_prot, pts_fib, pts_FLN]):
            return None
        return pts_A - sum([pts_prot, pts_fib, pts_FLN])


class ZugesetzteFette(NurtiScoreGeneral):
    @classmethod
    def calculate_nutri_score(
        cls,
        kilokalorien,
        gesaettigte_fettsaeuren,
        zucker,
        proteine,
        ballaststoffe,
        anteil_obst_gemuese_huelsen_schalen_raps_walnuss_und_olivenoele,
        salz=None,
        natrium=None,
        gesamtfett=None,
        ist_wasser=False,
    ):

        if salz == None and natrium == None:
            raise Exception("Salz oder Natrium muss angegeben werden")
        elif not salz == None:
            natrium_calculated = cls._calculate_natrium(salz)
            if not natrium == None:
                if natrium_calculated != natrium:
                    raise Exception(
                        "Entweder Salz und Natrium weichen stimmen nicht überein"
                    )
            natrium = natrium_calculated

        brennwert = cls._calculate_brennwert(kilokalorien)
        pts_kj = cls._calculate_pts_kj(brennwert)
        pts_glus = cls._calculate_pts_glus(zucker)
        pts_ags_lip_tot = cls._calculate_ags_lip_tot(
            gesaettigte_fettsaeuren, gesamtfett
        )
        pts_ags = cls._calculate_pts_ags(pts_ags_lip_tot)
        pts_na = cls._calculate_pts_na(natrium)
        pts_prot = cls._calculate_pts_prot(proteine)
        pts_fib = cls._calculate_pts_fib(ballaststoffe)
        pts_FLN = cls._calculate_pts_FLN(
            anteil_obst_gemuese_huelsen_schalen_raps_walnuss_und_olivenoele
        )
        pts_A = cls._calculate_pts_A(
            pts_kj,
            pts_glus,
            pts_ags,
            pts_na,
        )

        wert = cls._calculate_wert(pts_A, pts_prot, pts_fib, pts_FLN)
        rating = cls._calculate_rating(ist_wasser, wert)
        return rating, wert

    @staticmethod
    @_all_parameter_set
    def _calculate_ags_lip_tot(gesaettigte_fettsaeuren, gesamtfett):
        return round(gesaettigte_fettsaeuren / gesamtfett * 100, 2)

    @staticmethod
    @_all_parameter_set
    def _calculate_pts_ags(ags_lip_tot):
        wert = round(ags_lip_tot, 1)
        thresholds = {
            10: 0,
            16: 1,
            22: 2,
            28: 3,
            34: 4,
            40: 5,
            46: 6,
            52: 7,
            58: 8,
            64: 9,
        }
        return next((val for key, val in thresholds.items() if wert < key), 10)

    @staticmethod
    @_all_parameter_set
    def _calculate_wert(pts_A, pts_prot, pts_fib, pts_FLN):
        if pts_A < 11:
            return pts_A - (pts_prot + pts_fib + pts_FLN)
        elif pts_A >= 11 and pts_FLN == 5:
            return pts_A - (pts_prot + pts_fib + pts_FLN)
        else:
            return pts_A - pts_fib - pts_FLN


class Getraenke(NurtiScoreGeneral):
    @staticmethod
    @_all_parameter_set
    def _calculate_pts_kj(brennwert):
        wert = round(brennwert, 1)
        thresholds = {
            0: 0,
            30: 1,
            60: 2,
            90: 3,
            120: 4,
            150: 5,
            180: 6,
            210: 7,
            240: 8,
            270: 9,
        }
        return next((val for key, val in thresholds.items() if wert <= key), 10)

    @staticmethod
    @_all_parameter_set
    def _calculate_pts_glus(zucker):
        wert = round(zucker, 1)
        thresholds = {
            0: 0,
            1.5: 1,
            3: 2,
            4.5: 3,
            6: 4,
            7.5: 5,
            9: 6,
            10.5: 7,
            12: 8,
            13.5: 9,
        }
        return next((val for key, val in thresholds.items() if wert <= key), 10)

    @staticmethod
    @_all_parameter_set
    def _calculate_pts_FLN(
        anteil_obst_gemuese_huelsen_schalen_raps_walnuss_und_olivenoele,
    ):
        anteile = anteil_obst_gemuese_huelsen_schalen_raps_walnuss_und_olivenoele
        if anteile is None:
            return None
        wert = round(anteil_obst_gemuese_huelsen_schalen_raps_walnuss_und_olivenoele, 1)
        thresholds = {
            40: 0,
            60: 2,
            80: 4,
        }
        return next((val for key, val in thresholds.items() if wert <= key), 10)

    @staticmethod
    def _calculate_rating(ist_wasser, wert):
        if wert is None:
            return "fehlende_Daten"
        if wert < 0 or ist_wasser:
            return "Nutriscore_A"
        elif wert < 2:
            return "Nutriscore_B"
        elif wert < 6:
            return "Nutriscore_C"
        elif wert < 10:
            return "Nutriscore_D"
        else:
            return "Nutriscore_E"


class NutriScoreCategory(Enum):
    ALLGEMEINER_FALL = NurtiScoreGeneral
    KAESE = Kaese
    ZUGESETZTE_FETTE = ZugesetzteFette
    GETRAENKE = Getraenke


class NutriScoreGenerator:
    """
    Dieser Nutri-Score Rechner nutzt die Formeln aus der Nutri-Score Berechnungs Excel Tablle des BMEL von: https://www.bmel.de/SharedDocs/Downloads/DE/_Ernaehrung/Lebensmittel-Kennzeichnung/nutri-score-dt-excel-berechnungstabelle.html bezogen am 01.01.2024.
    Lediglich der Allgemeine Fall wird ausgerechnet.
    """

    @staticmethod
    def calculate_nutri_score(
        category: NutriScoreCategory,
        kilokalorien,
        gesaettigte_fettsaeuren,
        zucker,
        proteine,
        ballaststoffe,
        anteil_obst_gemuese_huelsen_schalen_raps_walnuss_und_olivenoele,
        salz=None,
        natrium=None,
        gesamtfett=None,
        ist_wasser=False,
    ):
        return category.value().calculate_nutri_score(
            kilokalorien=kilokalorien,
            gesaettigte_fettsaeuren=gesaettigte_fettsaeuren,
            zucker=zucker,
            proteine=proteine,
            ballaststoffe=ballaststoffe,
            anteil_obst_gemuese_huelsen_schalen_raps_walnuss_und_olivenoele=anteil_obst_gemuese_huelsen_schalen_raps_walnuss_und_olivenoele,
            salz=salz,
            natrium=natrium,
            gesamtfett=gesamtfett,
            ist_wasser=ist_wasser,
        )

def _extract_formulars():
    import openpyxl

    workbook = openpyxl.load_workbook(
        "./nutri-score-dt-excel-berechnungstabelle.xlsx",
        read_only=True,
    )
    sheetnames = ["allgemeiner Fall", "Käse", "zugesetzte Fette", "Getränke"]
    sheet_formulars = {}
    for sheetname in sheetnames:
        worksheet = workbook[sheetname]
        header = worksheet[1]
        columns = {col.column_letter: col.value for col in header}
        formulars = []
        for col, name in columns.items():
            cell = col + "2"
            formulars.append(f"{cell}, {name}: {worksheet[cell].value}")
        sheet_formulars[sheetname] = formulars
    workbook.close()
    import json

    with open("formulars.json", "w", encoding="utf-8") as f:
        json.dump(sheet_formulars, f, indent=6, ensure_ascii=False)


if __name__ == "__main__":
    _extract_formulars()
    pass
